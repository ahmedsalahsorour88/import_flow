"""
CGX-004: Customs Packing List Excel Generator.
Generates authentic Customs Packing List (.xlsx) workbooks for Egyptian Customs.
Supports 4 structure modes:
- by_hs_code:  one row per HS Code (most common in customs)
- flat:        one row per line item (full detail)
- by_pallet:   one section per pallet with items breakdown
- by_carton:   one row per carton/package
"""

import io
from datetime import datetime
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from modules.cargox.schemas import PackingListPayload, PackingListLineItem


# ── Color Palette (AGENTS.md colors) ──────────────────────────────────────────
_EMERALD_DARK = "1A7A40"      # header background
_EMERALD_LIGHT = "E8F5E9"     # alternating row tint
_CHARCOAL = "2C3E50"          # title text
_COBALT = "3498DB"            # sub-header
_CLOUD = "ECF0F1"             # label cell
_ORANGE = "E67E22"            # totals row
_WHITE = "FFFFFF"
_BLACK = "000000"


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(bold: bool = True, size: int = 10, color: str = _WHITE) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _normal_font(bold: bool = False, size: int = 9) -> Font:
    return Font(name="Calibri", bold=bold, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _center_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _merge_and_set(ws, cell_range: str, value: str, font: Font, fill: PatternFill, align: Alignment):
    ws.merge_cells(cell_range)
    first_col_letter = cell_range.split(":")[0].rstrip("0123456789")
    first_row = int("".join(filter(str.isdigit, cell_range.split(":")[0])))
    cell = ws[f"{first_col_letter}{first_row}"]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = align


def _write_meta_row(ws, row: int, label: str, value: str, col_start: int = 1):
    lbl_cell = ws.cell(row=row, column=col_start, value=label)
    lbl_cell.font = Font(name="Calibri", bold=True, size=9)
    lbl_cell.fill = _fill(_CLOUD)
    lbl_cell.alignment = _left_align(wrap=False)
    lbl_cell.border = _thin_border()

    val_cell = ws.cell(row=row, column=col_start + 1, value=value)
    val_cell.font = Font(name="Calibri", size=9)
    val_cell.alignment = _left_align(wrap=False)
    val_cell.border = _thin_border()


def generate_packing_list_excel_bytes(
    payload: PackingListPayload,
    structure: str = "by_hs_code",
    track_code: str = "",
    mode_label: str = "",
) -> bytes:
    """
    يُولّد شيت إكسل رسمي لقائمة التعبئة الجمركية.

    Args:
        payload:      PackingListPayload الكامل.
        structure:    "by_hs_code" | "flat" | "by_pallet" | "by_carton"
        track_code:   رمز المسار الجمركي (للرأسية).
        mode_label:   نص وصفي لـ mode يظهر في الرأسية.
    Returns:
        bytes of .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customs Packing List"

    # ── Column Widths ──────────────────────────────────────────────────────────
    col_widths = [6, 14, 30, 20, 10, 8, 12, 12, 14]
    col_headers_en = ["#", "HS Code", "Description", "Manufacturer", "Qty", "Unit",
                      "Net Wt (kg)", "Gross Wt (kg)", "Package Ref"]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25

    total_cols = len(col_headers_en)
    last_col = get_column_letter(total_cols)

    # ── Banner Row 1: Title ────────────────────────────────────────────────────
    _merge_and_set(
        ws, f"A1:{last_col}1",
        "CUSTOMS PACKING LIST  —  قائمة التعبئة الجمركية",
        Font(name="Calibri", bold=True, size=14, color=_WHITE),
        _fill(_EMERALD_DARK),
        _center_align(),
    )

    # ── Banner Row 2: Subtitle ─────────────────────────────────────────────────
    subtitle = f"{track_code or ''}  |  {mode_label or structure.upper()}  |  {datetime.now().strftime('%Y-%m-%d')}"
    _merge_and_set(
        ws, f"A2:{last_col}2",
        subtitle,
        Font(name="Calibri", bold=False, size=9, color=_WHITE),
        _fill(_COBALT),
        _center_align(),
    )

    # ── Metadata Block (rows 3–9) ──────────────────────────────────────────────
    meta_rows = [
        ("Seller / Shipper:", payload.seller_name or ""),
        ("Buyer / Importer:", payload.buyer_name or ""),
        ("ACID Number:", payload.acid_number or ""),
        ("Packing List Ref:", payload.packing_list_ref or "PL-001"),
        ("Invoice Number:", payload.invoice_number or ""),
        ("Port of Loading:", payload.origin_port or ""),
        ("Port of Discharge:", payload.destination_port or ""),
    ]
    meta_start_row = 3
    for offset, (lbl, val) in enumerate(meta_rows):
        _write_meta_row(ws, meta_start_row + offset, lbl, val)
        # merge value cell across 3 columns
        r = meta_start_row + offset
        ws.merge_cells(f"B{r}:{last_col}{r}")
        ws[f"B{r}"].font = Font(name="Calibri", size=9)
        ws[f"B{r}"].alignment = _left_align(wrap=False)

    header_row = meta_start_row + len(meta_rows) + 1  # row 11

    # ── Column Headers ─────────────────────────────────────────────────────────
    ws.row_dimensions[header_row].height = 20
    for col_idx, hdr in enumerate(col_headers_en, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=hdr)
        cell.font = _header_font(bold=True, size=9)
        cell.fill = _fill(_EMERALD_DARK)
        cell.alignment = _center_align()
        cell.border = _thin_border()

    # ── Data Rows ──────────────────────────────────────────────────────────────
    data_start = header_row + 1
    items_to_render = payload.items

    # Special rendering for by_pallet: group items by pallet_number
    if structure == "by_pallet" and items_to_render:
        pallets_map: Dict[str, List[PackingListLineItem]] = {}
        for itm in items_to_render:
            key = itm.pallet_number or itm.package_ref or "PLT-???"
            pallets_map.setdefault(key, []).append(itm)

        current_row = data_start
        row_count = 0
        for pallet_ref, pallet_items in pallets_map.items():
            # Pallet header sub-row
            sub_cell = ws.cell(row=current_row, column=1, value=pallet_ref)
            sub_cell.font = Font(name="Calibri", bold=True, size=9, color=_WHITE)
            sub_cell.fill = _fill(_COBALT)
            sub_cell.alignment = _center_align()
            ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
            ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=9, color=_WHITE)
            ws[f"A{current_row}"].fill = _fill(_COBALT)
            ws[f"A{current_row}"].alignment = _center_align()
            current_row += 1

            for itm in pallet_items:
                row_count += 1
                fill_color = _EMERALD_LIGHT if row_count % 2 == 0 else _WHITE
                _render_item_row(ws, current_row, row_count, itm, fill_color)
                current_row += 1

        data_end = current_row - 1
    else:
        for row_count, itm in enumerate(items_to_render, start=1):
            current_row = data_start + row_count - 1
            fill_color = _EMERALD_LIGHT if row_count % 2 == 0 else _WHITE
            _render_item_row(ws, current_row, row_count, itm, fill_color)
        data_end = data_start + len(items_to_render) - 1
        current_row = data_end + 1

    totals_row = current_row + 1

    # ── Totals Row ─────────────────────────────────────────────────────────────
    totals_data = [
        (1, "TOTALS"),
        (5, payload.total_packages),
        (7, payload.total_net_weight_kg),
        (8, payload.total_gross_weight_kg),
    ]
    ws.row_dimensions[totals_row].height = 18
    # fill entire totals row
    for c in range(1, total_cols + 1):
        tc = ws.cell(row=totals_row, column=c)
        tc.fill = _fill(_ORANGE)
        tc.border = _thin_border()
    for (col, val) in totals_data:
        tc = ws.cell(row=totals_row, column=col, value=val)
        tc.font = Font(name="Calibri", bold=True, size=9, color=_WHITE)
        tc.fill = _fill(_ORANGE)
        tc.alignment = _center_align()
        tc.border = _thin_border()

    # ── Signature Footer ───────────────────────────────────────────────────────
    sig_row = totals_row + 2
    sig_font = Font(name="Calibri", size=9)
    sig_labels = ["Prepared By:", "Verified By:", "Customs Officer:"]
    sig_cols = [1, 4, 7]
    for lbl, c in zip(sig_labels, sig_cols):
        lc = ws.cell(row=sig_row, column=c, value=lbl)
        lc.font = Font(name="Calibri", bold=True, size=9)
        ws.merge_cells(f"{get_column_letter(c)}{sig_row}:{get_column_letter(c+1)}{sig_row}")
        underline_row = sig_row + 1
        for uc in range(c, c + 2):
            uc_cell = ws.cell(row=underline_row, column=uc)
            uc_cell.border = Border(bottom=Side(style="thin", color="000000"))

    # freeze header
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _render_item_row(
    ws,
    row: int,
    row_num: int,
    itm: PackingListLineItem,
    fill_color: str,
):
    """كتابة سطر بند واحد في شيت الباكينج ليست."""
    row_fill = _fill(fill_color)
    border = _thin_border()

    cells_data = [
        (1, row_num, _center_align()),
        (2, itm.hs_code or "", _center_align()),
        (3, itm.description or "", _left_align()),
        (4, itm.manufacturer or "", _center_align()),
        (5, round(itm.quantity, 2), _center_align()),
        (6, itm.qty_unit or "PCS", _center_align()),
        (7, round(itm.net_weight_kg, 2), _center_align()),
        (8, round(itm.gross_weight_kg, 2), _center_align()),
        (9, itm.package_ref or "", _center_align()),
    ]
    for (col, val, align) in cells_data:
        c = ws.cell(row=row, column=col, value=val)
        c.font = _normal_font(size=9)
        c.fill = row_fill
        c.alignment = align
        c.border = border
