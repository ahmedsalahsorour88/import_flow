from datetime import datetime, date
from typing import Optional, Dict, Any, List
import io
import csv
from sqlalchemy.orm import Session
from fastapi import HTTPException
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .model import LandedCostSettlementRecord
from .schemas import (
    OdooJournalLineItem,
    OdooJournalEntryResponse,
    OdooExportConfig,
)
from modules.import_files.model import ImportFile

def generate_odoo_journal_entry_service(
    db: Session,
    settlement_id: int,
    config: Optional[OdooExportConfig] = None,
) -> OdooJournalEntryResponse:
    """
    Generates a balanced, double-entry Journal Entry for Odoo / ERP systems.
    Captures:
    1. Foreign Supplier Goods (FOB)
    2. Actual Freight (Sea/Air/Land)
    3. Customs Clearance Brokerage Invoice
    4. Inland Transport Invoice
    5. Customs Duty & Official Taxes Claim
    6. Demurrage & Storage Expenses
    7. Price Adjustment Penalties & Other Import Costs
    """
    if config is None:
        config = OdooExportConfig()

    settlement = db.query(LandedCostSettlementRecord).filter(
        LandedCostSettlementRecord.settlement_id == settlement_id,
        LandedCostSettlementRecord.is_active == True,
    ).first()

    if not settlement:
        raise HTTPException(status_code=404, detail="سجل التسوية المالية غير موجود.")

    imp_file = db.query(ImportFile).filter(
        ImportFile.import_file_id == settlement.import_file_id
    ).first()

    file_code = imp_file.import_file_code if imp_file else f"IMP-FILE-{settlement.import_file_id}"
    company_name = imp_file.company_name if imp_file else "الشركة المستوردة"
    supplier_name = imp_file.supplier_name if imp_file else "المورد الأجنبي"
    project_name = getattr(imp_file, "project_names", None) or getattr(imp_file, "project_name", None) or f"Project-{file_code}"

    entry_date = settlement.created_at.strftime("%Y-%m-%d") if settlement.created_at else date.today().strftime("%Y-%m-%d")
    ref_string = f"{file_code} / {settlement.settlement_code}"

    lines: List[OdooJournalLineItem] = []

    # 1. DEBIT: Total Landed Cost -> Goods in Transit / Inventory Valuation Account
    total_landed_cost = round(settlement.total_landed_cost_egp, 2)
    lines.append(
        OdooJournalLineItem(
            account_code=config.inventory_account_code,
            account_name=config.inventory_account_name,
            partner_name=company_name,
            label=f"إثبات تكلفة وصول الشحنة الإجمالية (Landed Cost) - ملف {file_code}",
            debit=total_landed_cost,
            credit=0.0,
            currency="EGP",
            analytic_account=project_name,
            cost_category="Goods",
        )
    )

    # 2. CREDIT: Foreign Supplier Payable (FOB Value)
    total_fob_egp = round(settlement.total_fob_egp, 2)
    if total_fob_egp > 0:
        lines.append(
            OdooJournalLineItem(
                account_code=config.supplier_account_code,
                account_name=f"موردون خارجيون - {supplier_name}",
                partner_name=supplier_name,
                label=f"قيمة بضاعة الفاتورة التجارية (FOB Goods) - المورد {supplier_name}",
                debit=0.0,
                credit=total_fob_egp,
                currency="EGP",
                analytic_account=project_name,
                cost_category="Goods",
            )
        )

    # 3. CREDIT: Breakdown of Actual Expense Invoices
    # Categorizes freight, customs, clearance, transport, demurrage, and adjustments
    expense_invoices = settlement.expense_invoices or []
    for exp in expense_invoices:
        category = exp.get("category", "Other").strip()
        provider = exp.get("provider_name", "مقدم الخدمة").strip()
        inv_no = exp.get("invoice_no", "N/A").strip()
        amount_egp = round(exp.get("amount_egp", 0.0), 2)
        if amount_egp <= 0:
            amount_egp = round(exp.get("amount_fx", 0.0) * exp.get("exchange_rate", 1.0), 2)

        if amount_egp <= 0:
            continue

        cat_lower = category.lower()
        if "freight" in cat_lower or "shipping" in cat_lower or "نولون" in cat_lower:
            acct_code = config.freight_account_code
            acct_name = f"موردو خدمات الشحن والنولون - {provider}"
            cost_cat = "Freight"
            desc_label = f"فاتورة نولون شحن رقم {inv_no} - شركة {provider}"
        elif "customs" in cat_lower or "duty" in cat_lower or "جمارك" in cat_lower or "ضرائب" in cat_lower or "vat" in cat_lower:
            acct_code = config.customs_authority_account_code
            acct_name = f"مصلحة الجمارك والضرائب - {provider}"
            cost_cat = "Customs"
            desc_label = f"مطالبة جمركية ورسوم إقرار 46 رقم {inv_no} - {provider}"
        elif "broker" in cat_lower or "clearance" in cat_lower or "تخليص" in cat_lower:
            acct_code = config.customs_broker_account_code
            acct_name = f"مستخلصو الجمارك والأتعاب - {provider}"
            cost_cat = "Clearance"
            desc_label = f"أتعاب ومصاريف تخليص جمركي فاتورة {inv_no} - {provider}"
        elif "transport" in cat_lower or "truck" in cat_lower or "نقل" in cat_lower:
            acct_code = config.transport_account_code
            acct_name = f"مقاولو النقل الداخلي - {provider}"
            cost_cat = "Transport"
            desc_label = f"فاتورة نقل داخلي وتعتيق رقم {inv_no} - {provider}"
        elif "demurrage" in cat_lower or "storage" in cat_lower or "ارضيات" in cat_lower or "أرضيات" in cat_lower or "غرامات" in cat_lower:
            acct_code = config.demurrage_account_code
            acct_name = f"غرامات أرضيات وتأخير حاويات - {provider}"
            cost_cat = "Demurrage"
            desc_label = f"غرامات أرضيات وتأخير حاويات فاتورة {inv_no} - {provider}"
        elif "adjustment" in cat_lower or "price" in cat_lower or "تعديل" in cat_lower or "فرق" in cat_lower:
            acct_code = config.price_adjustment_account_code
            acct_name = f"فروق وغرامات تسوية أسعار - {provider}"
            cost_cat = "Price_Adjustment"
            desc_label = f"غرامة وفروق تسوية قيمة استيرادية رقم {inv_no} - {provider}"
        else:
            acct_code = config.other_expenses_account_code
            acct_name = f"مصروفات استيراد متنوعة - {provider}"
            cost_cat = "Other"
            desc_label = f"مصروفات استيرادية {category} فاتورة {inv_no} - {provider}"

        lines.append(
            OdooJournalLineItem(
                account_code=acct_code,
                account_name=acct_name,
                partner_name=provider,
                label=desc_label,
                debit=0.0,
                credit=amount_egp,
                currency="EGP",
                amount_currency=round(exp.get("amount_fx", 0.0), 2) if exp.get("currency") != "EGP" else None,
                analytic_account=project_name,
                cost_category=cost_cat,
            )
        )

    # Balance Calculation
    total_debit = round(sum(l.debit for l in lines), 2)
    total_credit = round(sum(l.credit for l in lines), 2)
    diff = round(abs(total_debit - total_credit), 2)
    is_balanced = (diff <= 0.05)

    return OdooJournalEntryResponse(
        settlement_id=settlement.settlement_id,
        settlement_code=settlement.settlement_code,
        import_file_code=file_code,
        company_name=company_name,
        supplier_name=supplier_name,
        project_name=project_name,
        entry_date=entry_date,
        journal_name="Miscellaneous Operations / Vendor Bills",
        reference=ref_string,
        total_debit=total_debit,
        total_credit=total_credit,
        is_balanced=is_balanced,
        difference=diff,
        lines=lines,
        items_breakdown=settlement.item_landed_costs or [],
    )

def export_odoo_csv_service(db: Session, settlement_id: int) -> str:
    """
    Exports a clean CSV file strictly formatted for Odoo 15/16/17/18 `account.move` and `account.move.line` import.
    """
    entry = generate_odoo_journal_entry_service(db, settlement_id)
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # Standard Odoo Move Import Header
    writer.writerow([
        "id",
        "date",
        "ref",
        "journal_id/code",
        "line_ids/account_id/code",
        "line_ids/partner_id/name",
        "line_ids/name",
        "line_ids/debit",
        "line_ids/credit",
        "line_ids/currency_id/name",
        "line_ids/amount_currency",
        "line_ids/analytic_distribution",
    ])

    move_id_ref = f"importflow_settlement_{entry.settlement_id}"

    for idx, line in enumerate(entry.lines):
        analytic_str = f"{entry.project_name}" if entry.project_name else ""
        writer.writerow([
            move_id_ref if idx == 0 else "",
            entry.entry_date if idx == 0 else "",
            entry.reference if idx == 0 else "",
            "MISC" if idx == 0 else "",
            line.account_code,
            line.partner_name,
            line.label,
            f"{line.debit:.2f}" if line.debit > 0 else "0.00",
            f"{line.credit:.2f}" if line.credit > 0 else "0.00",
            line.currency,
            f"{line.amount_currency:.2f}" if line.amount_currency else "",
            analytic_str,
        ])

    return output.getvalue()

def export_odoo_excel_service(db: Session, settlement_id: int) -> bytes:
    """
    Generates a beautifully formatted, bilingual Excel Workbook for Accounting and Audit.
    Contains:
    - Sheet 1: Odoo Import Format (account.move)
    - Sheet 2: Accounting Journal Voucher (قيد اليومية المزدوج التفصيلي)
    - Sheet 3: Item Landed Cost Breakdown (تكلفة وصول الأصناف)
    """
    entry = generate_odoo_journal_entry_service(db, settlement_id)
    wb = openpyxl.Workbook()

    # Define Colors & Styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Flat Charcoal
    debit_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # Soft Emerald
    credit_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # Soft Amber
    total_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid") # Flat Cobalt

    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    font_total = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    thick_bottom = Border(bottom=Side(style="double", color="000000"), top=Side(style="thin", color="000000"))

    # ==========================================
    # SHEET 1: ODOO IMPORT SHEET
    # ==========================================
    ws_odoo = wb.active
    ws_odoo.title = "Odoo_GL_Import"

    headers_odoo = [
        "id", "date", "ref", "journal_id/code",
        "line_ids/account_id/code", "line_ids/partner_id/name",
        "line_ids/name", "line_ids/debit", "line_ids/credit",
        "line_ids/currency_id/name", "line_ids/amount_currency",
        "line_ids/analytic_distribution"
    ]
    ws_odoo.append(headers_odoo)
    for col_idx in range(1, len(headers_odoo) + 1):
        cell = ws_odoo.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    move_id_ref = f"importflow_settlement_{entry.settlement_id}"
    for idx, line in enumerate(entry.lines):
        ws_odoo.append([
            move_id_ref if idx == 0 else "",
            entry.entry_date if idx == 0 else "",
            entry.reference if idx == 0 else "",
            "MISC" if idx == 0 else "",
            line.account_code,
            line.partner_name,
            line.label,
            line.debit,
            line.credit,
            line.currency,
            line.amount_currency or "",
            entry.project_name or "",
        ])
        row_idx = idx + 2
        for c in range(1, len(headers_odoo) + 1):
            c_cell = ws_odoo.cell(row=row_idx, column=c)
            c_cell.font = font_regular
            c_cell.border = thin_border
            if c in (8, 9):
                c_cell.number_format = "#,##0.00"

    # ==========================================
    # SHEET 2: ACCOUNTING JOURNAL VOUCHER (سند قيد اليومية)
    # ==========================================
    ws_voucher = wb.create_sheet(title="Accounting_Voucher")
    ws_voucher.views.sheetView[0].rightToLeft = True

    # Title Block
    ws_voucher.merge_cells("A1:G1")
    title_cell = ws_voucher.cell(row=1, column=1, value=f"سند قيد إقفال وتكلفة وصول الاستيراد (Import Landed Cost Journal Voucher)")
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_voucher.row_dimensions[1].height = 30

    # Meta Info Box
    meta_info = [
        ("رقم ملف الشحنة:", entry.import_file_code, "رقم قيد التسوية:", entry.settlement_code),
        ("الشركة المستوردة:", entry.company_name, "المورد الأجنبي:", entry.supplier_name),
        ("تاريخ القيد:", entry.entry_date, "المشروع / مركز التكلفة:", entry.project_name or "N/A"),
    ]
    for r_i, (k1, v1, k2, v2) in enumerate(meta_info, start=3):
        ws_voucher.cell(row=r_i, column=1, value=k1).font = font_bold
        ws_voucher.cell(row=r_i, column=2, value=v1).font = font_regular
        ws_voucher.cell(row=r_i, column=4, value=k2).font = font_bold
        ws_voucher.cell(row=r_i, column=5, value=v2).font = font_regular

    # Table Headers
    t_headers = ["رقم الحساب", "اسم الحساب الدفتري", "الجهة / الطرف (Partner)", "شرح وتفاصيل القيد (Label)", "مدين (Debit EGP)", "دائن (Credit EGP)", "تصنيف المصروف"]
    start_r = 7
    for c_i, h in enumerate(t_headers, start=1):
        c = ws_voucher.cell(row=start_r, column=c_i, value=h)
        c.fill = header_fill
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_voucher.row_dimensions[start_r].height = 24

    cur_r = start_r + 1
    for line in entry.lines:
        ws_voucher.cell(row=cur_r, column=1, value=line.account_code).alignment = Alignment(horizontal="center")
        ws_voucher.cell(row=cur_r, column=2, value=line.account_name)
        ws_voucher.cell(row=cur_r, column=3, value=line.partner_name)
        ws_voucher.cell(row=cur_r, column=4, value=line.label)

        d_cell = ws_voucher.cell(row=cur_r, column=5, value=line.debit if line.debit > 0 else "")
        d_cell.number_format = "#,##0.00"
        if line.debit > 0:
            d_cell.fill = debit_fill

        c_cell = ws_voucher.cell(row=cur_r, column=6, value=line.credit if line.credit > 0 else "")
        c_cell.number_format = "#,##0.00"
        if line.credit > 0:
            c_cell.fill = credit_fill

        ws_voucher.cell(row=cur_r, column=7, value=line.cost_category).alignment = Alignment(horizontal="center")

        for col_x in range(1, 8):
            ws_voucher.cell(row=cur_r, column=col_x).border = thin_border
            ws_voucher.cell(row=cur_r, column=col_x).font = font_regular

        cur_r += 1

    # Totals Row
    ws_voucher.cell(row=cur_r, column=1, value="الإجمالي العام والمتوازن").font = font_total
    ws_voucher.cell(row=cur_r, column=1).fill = total_fill
    ws_voucher.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=4)

    tot_d = ws_voucher.cell(row=cur_r, column=5, value=entry.total_debit)
    tot_d.font = font_total
    tot_d.fill = total_fill
    tot_d.number_format = "#,##0.00"
    tot_d.border = thick_bottom

    tot_c = ws_voucher.cell(row=cur_r, column=6, value=entry.total_credit)
    tot_c.font = font_total
    tot_c.fill = total_fill
    tot_c.number_format = "#,##0.00"
    tot_c.border = thick_bottom

    status_str = "🟢 متوازن 100%" if entry.is_balanced else f"🔴 غير متوازن (فرق: {entry.difference:.2f})"
    ws_voucher.cell(row=cur_r, column=7, value=status_str).font = font_total
    ws_voucher.cell(row=cur_r, column=7).fill = total_fill
    ws_voucher.cell(row=cur_r, column=7).alignment = Alignment(horizontal="center")

    # ==========================================
    # SHEET 3: ITEM LANDED COSTS BREAKDOWN
    # ==========================================
    ws_items = wb.create_sheet(title="Items_Landed_Cost")
    ws_items.views.sheetView[0].rightToLeft = True

    i_headers = [
        "كود الصنف", "اسم الصنف", "الكمية", "تكلفة FOB للوحدة", "إجمالي FOB",
        "نصيب النولون", "نصيب الجمارك", "نصيب التخليص", "نصيب النقل", "مصروفات أخرى",
        "إجمالي تكلفة الصنف", "تكلفة الوحدة النهائية", "معامل الزيادة (Markup)"
    ]
    ws_items.append(i_headers)
    for c_idx in range(1, len(i_headers) + 1):
        cell = ws_items.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, itm in enumerate(entry.items_breakdown, start=2):
        ws_items.append([
            itm.get("item_code", ""),
            itm.get("item_name", ""),
            itm.get("qty", 1),
            itm.get("fob_unit_egp", 0.0),
            itm.get("fob_total_egp", 0.0),
            itm.get("allocated_freight_egp", 0.0),
            itm.get("allocated_customs_egp", 0.0),
            itm.get("allocated_clearance_egp", 0.0),
            itm.get("allocated_transport_egp", 0.0),
            itm.get("allocated_other_egp", 0.0),
            itm.get("total_landed_cost_egp", 0.0),
            itm.get("unit_landed_cost_egp", 0.0),
            itm.get("markup_factor", 1.0),
        ])
        for c_x in range(1, len(i_headers) + 1):
            cell = ws_items.cell(row=idx, column=c_x)
            cell.font = font_regular
            cell.border = thin_border
            if c_x in range(4, 13):
                cell.number_format = "#,##0.00"

    # Auto-adjust column widths on all sheets
    for ws in [ws_odoo, ws_voucher, ws_items]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
