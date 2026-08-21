import io
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from fastapi import HTTPException, status
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.import_files.model import ImportFile
from modules.purchase_orders.model import PurchaseOrder
from modules.import_documentation.service import get_central_archive_service
from .model import OriginalDocumentsCollectionSession
from .repository import OriginalDocumentsCollectionRepository
from .schemas import (
    OriginalDocumentsCollectionCreate,
    OriginalDocumentsCollectionUpdate,
    OriginalDocumentsCollectionResponse,
    OriginalDocumentsAutoPopulateResponse,
    CourierEntry,
    OriginalDocumentItem,
)


class OriginalDocumentsCollectionService:
    @staticmethod
    def auto_populate_from_central_archive(
        db: Session, import_file_id: int
    ) -> OriginalDocumentsAutoPopulateResponse:
        import_file = db.query(ImportFile).filter_by(import_file_id=import_file_id).first()
        if not import_file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Import File ID {import_file_id} not found.",
            )

        existing_session = OriginalDocumentsCollectionRepository.get_by_import_file(db, import_file_id)

        # Retrieve central archive details if available
        archive_data = None
        try:
            archive_data = get_central_archive_service(db, import_file_id)
        except Exception:
            archive_data = None

        # Build baseline required documents list
        # Check inspection requirement from archive
        is_inspection_required = "No"
        if archive_data and archive_data.inspection_certificate.is_available:
            is_inspection_required = "Yes"
        elif archive_data and archive_data.goeic_inspection_alert:
            is_inspection_required = "Conditional"

        # Check COO requirement
        coo_required = "Yes"
        if archive_data and archive_data.certificate_of_origin.is_waived:
            coo_required = "No"

        required_docs = [
            OriginalDocumentItem(
                category="Commercial",
                document_name="Proforma Invoice",
                is_required="Yes",
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Commercial",
                document_name="Commercial Invoice",
                is_required="Yes",
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Commercial",
                document_name="Packing List",
                is_required="Yes",
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Commercial",
                document_name="E.Invoice",
                is_required="Yes",
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Certificate",
                document_name="Certificate of Origin (COO)",
                is_required=coo_required,
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Shipping",
                document_name="Final House Bill of Lading",
                is_required="Yes",
                responsible_party="Freight Forwarder",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Shipping",
                document_name="Final Master Bill of Lading",
                is_required="Yes",
                responsible_party="Freight Forwarder",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Certificate",
                document_name="Inspection Certificate",
                is_required=is_inspection_required,
                responsible_party="Supplier",
                status="Pending",
            ),
            OriginalDocumentItem(
                category="Egypt Import",
                document_name="Blockchain Documents",
                is_required="Conditional",
                responsible_party="Supplier",
                status="Pending",
            ),
        ]

        # Default Couriers
        default_couriers = [
            CourierEntry(
                courier_no="",
                courier_company="DHL",
                dispatch_date=None,
                is_received=False,
                received_date=None,
                received_by=None,
                notes="Foreign Exporter Hard-copy Package",
            ),
            CourierEntry(
                courier_no="",
                courier_company="FedEx",
                dispatch_date=None,
                is_received=False,
                received_date=None,
                received_by=None,
                notes="Freight Forwarder Shipping Documents Package",
            ),
        ]

        existing_response = None
        if existing_session:
            existing_response = OriginalDocumentsCollectionResponse.model_validate(existing_session)

        return OriginalDocumentsAutoPopulateResponse(
            import_file_id=import_file.import_file_id,
            import_file_code=import_file.import_file_code,
            acid_number=import_file.acid_number,
            importer_name=import_file.company_name if hasattr(import_file, "company_name") else getattr(import_file, "importer_name", ""),
            supplier_name=import_file.supplier_name if hasattr(import_file, "supplier_name") else "",
            default_couriers=default_couriers,
            required_documents=required_docs,
            existing_session=existing_response,
        )

    @staticmethod
    def save_or_upsert_collection_session(
        db: Session,
        payload: OriginalDocumentsCollectionCreate,
        username: str = "ADMIN",
    ) -> OriginalDocumentsCollectionResponse:
        import_file = db.query(ImportFile).filter_by(import_file_id=payload.import_file_id).first()
        if not import_file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Import File ID {payload.import_file_id} not found.",
            )

        # Calculate statistics
        docs_list = [d.model_dump() if isinstance(d, OriginalDocumentItem) else d for d in payload.documents_list]
        couriers_list = [c.model_dump() if isinstance(c, CourierEntry) else c for c in payload.couriers_list]

        total_count = len(docs_list)
        received_count = sum(1 for d in docs_list if d.get("is_received", False))
        verified_count = sum(1 for d in docs_list if d.get("is_verified", False))
        pending_count = total_count - received_count
        completion_pct = round((verified_count / total_count * 100.0), 1) if total_count > 0 else 0.0

        # Automatic status resolution if not explicitly set to a customized status
        status_val = payload.status
        if status_val in ["DRAFT", "PARTIALLY_RECEIVED", "FULLY_RECEIVED", "FULLY_VERIFIED"]:
            if verified_count == total_count and total_count > 0:
                status_val = "FULLY_VERIFIED"
            elif received_count == total_count and total_count > 0:
                status_val = "FULLY_RECEIVED"
            elif received_count > 0:
                status_val = "PARTIALLY_RECEIVED"
            else:
                status_val = "DRAFT"

        existing_session = OriginalDocumentsCollectionRepository.get_by_import_file(db, payload.import_file_id)

        if existing_session:
            # Update existing
            updates = {
                "acid_number": payload.acid_number or existing_session.acid_number,
                "importer_name": payload.importer_name or existing_session.importer_name,
                "supplier_name": payload.supplier_name or existing_session.supplier_name,
                "status": status_val,
                "couriers_list": couriers_list,
                "documents_list": docs_list,
                "total_documents_count": total_count,
                "received_documents_count": received_count,
                "verified_documents_count": verified_count,
                "pending_documents_count": pending_count,
                "completion_percentage": completion_pct,
                "discrepancy_override_reason": payload.discrepancy_override_reason,
                "notes": payload.notes,
                "updated_by": username,
            }
            saved = OriginalDocumentsCollectionRepository.update(db, existing_session, updates)
            return OriginalDocumentsCollectionResponse.model_validate(saved)
        else:
            # Create new
            new_code = OriginalDocumentsCollectionRepository.get_next_collection_code(db)
            new_session = OriginalDocumentsCollectionSession(
                collection_code=new_code,
                import_file_id=payload.import_file_id,
                import_file_code=payload.import_file_code,
                acid_number=payload.acid_number or import_file.acid_number,
                importer_name=payload.importer_name or getattr(import_file, "company_name", ""),
                supplier_name=payload.supplier_name or getattr(import_file, "supplier_name", ""),
                status=status_val,
                couriers_list=couriers_list,
                documents_list=docs_list,
                total_documents_count=total_count,
                received_documents_count=received_count,
                verified_documents_count=verified_count,
                pending_documents_count=pending_count,
                completion_percentage=completion_pct,
                discrepancy_override_reason=payload.discrepancy_override_reason,
                notes=payload.notes,
                created_by=username,
                updated_by=username,
            )
            saved = OriginalDocumentsCollectionRepository.create(db, new_session)
            return OriginalDocumentsCollectionResponse.model_validate(saved)

    @staticmethod
    def get_session_by_id(db: Session, collection_id: int) -> OriginalDocumentsCollectionResponse:
        session = OriginalDocumentsCollectionRepository.get_by_id(db, collection_id)
        if not session:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collection session not found.")
        return OriginalDocumentsCollectionResponse.model_validate(session)

    @staticmethod
    def get_session_by_file(db: Session, import_file_id: int) -> Optional[OriginalDocumentsCollectionResponse]:
        session = OriginalDocumentsCollectionRepository.get_by_import_file(db, import_file_id)
        if not session:
            return None
        return OriginalDocumentsCollectionResponse.model_validate(session)

    @staticmethod
    def list_sessions(
        db: Session,
        include_inactive: bool = False,
        status: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 100,
        offset: int = 0,
    ) -> List[OriginalDocumentsCollectionResponse]:
        sessions = OriginalDocumentsCollectionRepository.get_all(
            db,
            include_inactive=include_inactive,
            status=status,
            search=search,
            limit=limit,
            offset=offset,
        )
        return [OriginalDocumentsCollectionResponse.model_validate(s) for s in sessions]

    @staticmethod
    def export_collection_to_excel_bytes(
        db: Session, import_file_id_or_session_id: int
    ) -> bytes:
        # Resolve session
        session = OriginalDocumentsCollectionRepository.get_by_id(db, import_file_id_or_session_id)
        if not session:
            session = OriginalDocumentsCollectionRepository.get_by_import_file(db, import_file_id_or_session_id)

        if not session:
            # Generate from auto-populated
            auto_data = OriginalDocumentsCollectionService.auto_populate_from_central_archive(
                db, import_file_id_or_session_id
            )
            couriers = [c.model_dump() for c in auto_data.default_couriers]
            docs = [d.model_dump() for d in auto_data.required_documents]
            file_code = auto_data.import_file_code
            acid_no = auto_data.acid_number or "—"
            importer = auto_data.importer_name or "—"
            supplier = auto_data.supplier_name or "—"
            status_text = "DRAFT"
        else:
            couriers = session.couriers_list or []
            docs = session.documents_list or []
            file_code = session.import_file_code
            acid_no = session.acid_number or "—"
            importer = session.importer_name or "—"
            supplier = session.supplier_name or "—"
            status_text = session.status

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Original Docs Collection"

        # Fonts & Styles
        font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Calibri", size=11, bold=True, color="2C3E50")
        font_table_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=9, bold=False, color="000000")
        font_bold = Font(name="Calibri", size=9, bold=True, color="000000")

        fill_navy = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        fill_cobalt = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        fill_light_grey = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
        fill_alt_row = PatternFill(start_color="FBFCFC", end_color="FBFCFC", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="BDC3C7"),
            right=Side(style="thin", color="BDC3C7"),
            top=Side(style="thin", color="BDC3C7"),
            bottom=Side(style="thin", color="BDC3C7"),
        )

        # Title Row
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = f"PHYSICAL ORIGINAL DOCUMENTS COLLECTION & COURIER TRACKING SHEET — {file_code}"
        title_cell.font = font_title
        title_cell.fill = fill_navy
        title_cell.alignment = align_center
        ws.row_dimensions[1].height = 32

        # Metadata Header
        meta_items = [
            ("Import File:", file_code, "ACID Number:", acid_no),
            ("Importer Company:", importer, "Foreign Exporter:", supplier),
            ("Collection Status:", status_text, "Generated At:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ]

        row_idx = 3
        for m1_lbl, m1_val, m2_lbl, m2_val in meta_items:
            ws.cell(row=row_idx, column=1, value=m1_lbl).font = font_subtitle
            ws.cell(row=row_idx, column=2, value=m1_val).font = font_bold
            ws.cell(row=row_idx, column=6, value=m2_lbl).font = font_subtitle
            ws.cell(row=row_idx, column=7, value=m2_val).font = font_bold
            row_idx += 1

        row_idx += 1

        # Couriers Summary Section
        ws.cell(row=row_idx, column=1, value="COURIER DISPATCH PACKAGES (الطرود البريدية السريعة)").font = font_subtitle
        row_idx += 1

        courier_hdrs = ["#", "Courier No (رقم البوليصة)", "Courier Company (الشركة)", "Dispatch Date (تاريخ الإرسال)", "Received (الاستلام)", "Received Date (تاريخ الاستلام)", "Received By (المستلم)", "Notes (ملاحظات)"]
        for c_idx, h in enumerate(courier_hdrs, start=1):
            c = ws.cell(row=row_idx, column=c_idx, value=h)
            c.font = font_table_hdr
            c.fill = fill_cobalt
            c.alignment = align_center
            c.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

        if not couriers:
            couriers = [{"courier_no": "—", "courier_company": "DHL", "dispatch_date": "—", "is_received": False, "received_date": "—", "received_by": "—", "notes": ""}]

        for idx, cr in enumerate(couriers, start=1):
            ws.cell(row=row_idx, column=1, value=idx).alignment = align_center
            ws.cell(row=row_idx, column=2, value=cr.get("courier_no") or "—").alignment = align_left
            ws.cell(row=row_idx, column=3, value=cr.get("courier_company") or "DHL").alignment = align_center
            ws.cell(row=row_idx, column=4, value=cr.get("dispatch_date") or "—").alignment = align_center
            ws.cell(row=row_idx, column=5, value="Yes" if cr.get("is_received") else "No").alignment = align_center
            ws.cell(row=row_idx, column=6, value=cr.get("received_date") or "—").alignment = align_center
            ws.cell(row=row_idx, column=7, value=cr.get("received_by") or "—").alignment = align_left
            ws.cell(row=row_idx, column=8, value=cr.get("notes") or "").alignment = align_left

            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = font_data
                cell.border = thin_border
                if idx % 2 == 0:
                    cell.fill = fill_alt_row
            row_idx += 1

        row_idx += 2

        # Document Items Table
        ws.cell(row=row_idx, column=1, value="ORIGINAL HARD-COPY DOCUMENTS AUDIT & VERIFICATION (مصفوفة تدقيق واستلام أصول المستندات)").font = font_subtitle
        row_idx += 1

        doc_headers = [
            "Courier No",
            "Courier Date",
            "Document Category",
            "Document Name",
            "Required",
            "Responsible Party",
            "Received",
            "Received Date",
            "Verified",
            "Verified By",
            "Verification Date",
            "Status",
            "Remarks",
        ]

        for c_idx, h in enumerate(doc_headers, start=1):
            c = ws.cell(row=row_idx, column=c_idx, value=h)
            c.font = font_table_hdr
            c.fill = fill_navy
            c.alignment = align_center
            c.border = thin_border
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1

        # Populate rows
        for d_idx, doc in enumerate(docs, start=1):
            # Find courier date if courier_no matches
            c_no = doc.get("courier_no") or ""
            c_date = ""
            for cr in couriers:
                if cr.get("courier_no") and cr.get("courier_no") == c_no:
                    c_date = cr.get("dispatch_date") or ""
                    break

            ws.cell(row=row_idx, column=1, value=c_no or "—").alignment = align_left
            ws.cell(row=row_idx, column=2, value=c_date or "—").alignment = align_center
            ws.cell(row=row_idx, column=3, value=doc.get("category") or "Commercial").alignment = align_center
            ws.cell(row=row_idx, column=4, value=doc.get("document_name") or "").alignment = align_left
            ws.cell(row=row_idx, column=5, value=doc.get("is_required") or "Yes").alignment = align_center
            ws.cell(row=row_idx, column=6, value=doc.get("responsible_party") or "Supplier").alignment = align_center
            ws.cell(row=row_idx, column=7, value="Yes" if doc.get("is_received") else "No").alignment = align_center
            ws.cell(row=row_idx, column=8, value=doc.get("received_date") or "—").alignment = align_center
            ws.cell(row=row_idx, column=9, value="Yes" if doc.get("is_verified") else "No").alignment = align_center
            ws.cell(row=row_idx, column=10, value=doc.get("verified_by") or "—").alignment = align_left
            ws.cell(row=row_idx, column=11, value=doc.get("verification_date") or "—").alignment = align_center
            ws.cell(row=row_idx, column=12, value=doc.get("status") or "Pending").alignment = align_center
            ws.cell(row=row_idx, column=13, value=doc.get("remarks") or "").alignment = align_left

            for col in range(1, 14):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = font_data
                cell.border = thin_border
                if d_idx % 2 == 0:
                    cell.fill = fill_alt_row
            row_idx += 1

        # Adjust column widths automatically
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
