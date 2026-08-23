"""
Smart Document Upload — Service
Orchestrates file parsing + field extraction for all ImportFlow modules.
Routes uploaded file to the correct extractor based on module_name.
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from modules.smart_document_upload.extractors.purchase_order import PurchaseOrderExtractor
from modules.smart_document_upload.extractors.import_file import ImportFileExtractor
from modules.smart_document_upload.extractors.cargo_shipping import CargoShippingExtractor
from modules.smart_document_upload.extractors.customs_clearance import CustomsClearanceExtractor
from modules.smart_document_upload.extractors.customs_broker_quotation import CustomsBrokerQuotationExtractor
from modules.smart_document_upload.extractors.freight_quotation import FreightQuotationExtractor
from modules.smart_document_upload.extractors.freight_booking import FreightBookingExtractor
from modules.smart_document_upload.extractors.other_extractors import (
    COOCertificateExtractor,
    InspectionCertificateExtractor,
    FinancialDocumentExtractor,
)
from modules.smart_document_upload.extractors.master_data_entity import MasterDataEntityExtractor
from modules.smart_document_upload.extractors.base_extractor import BaseExtractor
import modules.smart_document_upload.repository as repo

# Entity verification imports
from modules.suppliers.model import Supplier
from modules.import_companies.model import ImportCompany


logger = logging.getLogger(__name__)

# ─── Extractor Registry ───────────────────────────────────────────────────────

_EXTRACTOR_MAP: Dict[str, BaseExtractor] = {
    "purchase-order": PurchaseOrderExtractor(),
    "import-file": ImportFileExtractor(),
    "cargo-shipping": CargoShippingExtractor(),
    "customs-clearance": CustomsClearanceExtractor(),
    "clearance-quotation": CustomsBrokerQuotationExtractor(),
    "customs-broker-quotation": CustomsBrokerQuotationExtractor(),
    "freight-quotation": FreightQuotationExtractor(),
    "freight-booking": FreightBookingExtractor(),
    "coo-certificate": COOCertificateExtractor(),
    "inspection-certificate": InspectionCertificateExtractor(),
    "financial-document": FinancialDocumentExtractor(),
    "master-data-entity": MasterDataEntityExtractor(),
    "supplier-entity": MasterDataEntityExtractor(),
    "importer-entity": MasterDataEntityExtractor(),
    "partner-entity": MasterDataEntityExtractor(),
    "bank-entity": MasterDataEntityExtractor(),
    # Aliases
    "customs-consultation": ImportFileExtractor(),   # extracts invoice fields
    "warehouse-receiving": ImportFileExtractor(),    # same as invoice extraction
    "demurrage": FinancialDocumentExtractor(),       # demurrage is a financial doc
}


# ─── Text Extraction ──────────────────────────────────────────────────────────

def extract_raw_text_and_boxes(filename: str, content_bytes: bytes) -> Tuple[str, dict]:
    """
    Reuses the existing extraction engine from import_documentation.
    Supports PDF (spatial), Word, Excel, and text files.
    """
    try:
        from modules.import_documentation.service import (
            extract_text_and_boxes_from_uploaded_file,
        )
        return extract_text_and_boxes_from_uploaded_file(filename, content_bytes)
    except Exception as e:
        logger.warning(f"Primary extractor failed for '{filename}': {e}. Using fallback.")
        return _fallback_text_extraction(filename, content_bytes), {}


def _fallback_text_extraction(filename: str, content_bytes: bytes) -> str:
    """Minimal text extraction fallback."""
    lower = filename.lower()
    try:
        if lower.endswith(".pdf"):
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(content_bytes))
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        elif lower.endswith((".docx", ".doc")):
            import docx
            doc = docx.Document(io.BytesIO(content_bytes))
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        elif lower.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if row_vals:
                        parts.append(" | ".join(row_vals))
            return "\n".join(parts)
        else:
            return content_bytes.decode("utf-8", errors="ignore")
    except Exception as e:
        return f"[Extraction error: {e}]"


# ─── Entity Verification ──────────────────────────────────────────────────────

def verify_extracted_entities(db: Session, extracted_fields: Dict[str, Any]) -> Dict[str, Any]:
    """
    After field extraction, checks whether the identified supplier_name and
    importer_name (consignee) exist in the database.

    Returns a dict with verification results:
    {
        supplier_verified: bool | None,
        supplier_id: int | None,
        supplier_code: str | None,
        importer_verified: bool | None,
        importer_id: int | None,
        importer_code: str | None,
    }
    Uses case-insensitive LIKE search to handle minor name differences.
    """
    result: Dict[str, Any] = {
        "supplier_verified": None,
        "supplier_id": None,
        "supplier_code": None,
        "importer_verified": None,
        "importer_id": None,
        "importer_code": None,
    }

    # ── Supplier verification ────────────────────────────────────────────────
    supplier_name: Optional[str] = extracted_fields.get("supplier_name") or extracted_fields.get("shipper")
    if supplier_name and supplier_name.strip():
        name_clean = supplier_name.strip()
        try:
            from sqlalchemy import func as sa_func
            matched_supplier = (
                db.query(Supplier)
                .filter(
                    Supplier.is_active == True,
                    sa_func.lower(Supplier.company_name).contains(name_clean.lower()[:40])
                )
                .first()
            )
            if matched_supplier:
                result["supplier_verified"] = True
                result["supplier_id"] = matched_supplier.supplier_id
                result["supplier_code"] = matched_supplier.supplier_code
            else:
                result["supplier_verified"] = False
        except Exception as e:
            logger.warning(f"Supplier verification query failed: {e}")
            result["supplier_verified"] = False

    # ── Importer / Consignee verification ────────────────────────────────────
    importer_name: Optional[str] = (
        extracted_fields.get("importer_name")
        or extracted_fields.get("consignee")
        or extracted_fields.get("company_name")
    )
    if importer_name and importer_name.strip():
        name_clean = importer_name.strip()
        try:
            from sqlalchemy import func as sa_func
            matched_company = (
                db.query(ImportCompany)
                .filter(
                    ImportCompany.is_active == True,
                    sa_func.lower(ImportCompany.importer_name).contains(name_clean.lower()[:40])
                )
                .first()
            )
            if matched_company:
                result["importer_verified"] = True
                result["importer_id"] = matched_company.company_id
                result["importer_code"] = matched_company.importer_id  # business code field
            else:
                result["importer_verified"] = False
        except Exception as e:
            logger.warning(f"Importer verification query failed: {e}")
            result["importer_verified"] = False

    return result


# ─── Main Service Function ────────────────────────────────────────────────────

def parse_uploaded_document(
    db: Session,
    module_name: str,
    filename: str,
    file_type: str,
    content_bytes: bytes,
    save_session: bool = True,
    created_by: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Main orchestrator:
    1. Extracts raw text from uploaded file (PDF/Excel/Word)
    2. Routes to the appropriate module extractor
    3. Computes confidence score and missing fields
    4. Saves an upload session record
    5. Returns a structured response dict
    """
    file_size = len(content_bytes)

    # Step 1 — Extract raw text
    raw_text, spatial_boxes = extract_raw_text_and_boxes(filename, content_bytes)

    # Step 2 — Route to extractor
    extractor = _EXTRACTOR_MAP.get(module_name)
    if extractor is None:
        logger.warning(f"No extractor for module '{module_name}'. Using ImportFileExtractor as default.")
        extractor = ImportFileExtractor()

    # Step 3 — Extract fields
    try:
        if isinstance(extractor, MasterDataEntityExtractor):
            extracted_fields = extractor.extract(raw_text, spatial_boxes, module_name=module_name)
        else:
            extracted_fields = extractor.extract(raw_text, spatial_boxes)
    except Exception as e:
        logger.error(f"Extractor error for '{module_name}': {e}")
        extracted_fields = {}

    # Step 4 — Compute quality metrics
    confidence = extractor.compute_confidence(extracted_fields)
    missing = extractor.missing_required(extracted_fields)

    if confidence >= 0.8:
        status = "SUCCESS"
    elif confidence >= 0.4:
        status = "PARTIAL"
    else:
        status = "FAILED"

    notes = (
        f"Extracted {len([v for v in extracted_fields.values() if v is not None])} fields. "
        f"Confidence: {confidence:.0%}. "
        + (f"Missing: {', '.join(missing)}." if missing else "All required fields found.")
    )

    # Step 5 — Verify extracted entities against DB (supplier + importer)
    entity_verification = verify_extracted_entities(db, extracted_fields)

    # Step 6 — Save session
    session_id = None
    session_ref = None
    if save_session:
        try:
            session = repo.create_upload_session(
                db=db,
                module_name=module_name,
                filename=filename,
                file_type=file_type,
                file_size_bytes=file_size,
                extraction_status=status,
                confidence_score=confidence,
                extracted_fields=extracted_fields,
                missing_fields=missing,
                extraction_notes=notes,
                created_by=created_by,
            )
            session_id = session.id
            session_ref = session.session_ref
        except Exception as e:
            logger.warning(f"Could not save upload session: {e}")

    return {
        "session_id": session_id,
        "session_ref": session_ref,
        "module_name": module_name,
        "filename": filename,
        "file_type": file_type,
        "file_size_bytes": file_size,
        "extraction_status": status,
        "confidence_score": confidence,
        "extracted_fields": extracted_fields,
        "missing_fields": missing,
        "extraction_notes": notes,
        "raw_text_preview": raw_text[:500] if raw_text else None,
        **entity_verification,
    }


def parse_multiple_uploaded_documents(
    db: Session,
    module_name: str,
    files_data: List[Tuple[str, str, bytes]],
    save_session: bool = True,
    created_by: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Parses multiple documents (e.g. Commercial Invoice + Packing List) concurrently,
    extracts raw text from each file, combines the raw text and spatial boxes,
    runs the module extractor, and returns a unified combined response.
    """
    combined_texts = []
    combined_boxes = {}
    combined_filenames = []
    total_size = 0

    for filename, file_type, content_bytes in files_data:
        raw_text, spatial_boxes = extract_raw_text_and_boxes(filename, content_bytes)
        combined_texts.append(f"=== DOCUMENT: {filename} ===\n{raw_text}")
        combined_boxes.update(spatial_boxes)
        combined_filenames.append(filename)
        total_size += len(content_bytes)

    full_text = "\n\n".join(combined_texts)
    unified_filename = " + ".join(combined_filenames)
    primary_file_type = files_data[0][1] if files_data else "pdf"

    extractor = _EXTRACTOR_MAP.get(module_name)
    if extractor is None:
        extractor = ImportFileExtractor()

    try:
        if isinstance(extractor, MasterDataEntityExtractor):
            extracted_fields = extractor.extract(full_text, combined_boxes, module_name=module_name)
        else:
            extracted_fields = extractor.extract(full_text, combined_boxes)
    except Exception as e:
        logger.error(f"Multi-document extractor error for '{module_name}': {e}")
        extracted_fields = {}

    confidence = extractor.compute_confidence(extracted_fields)
    missing = extractor.missing_required(extracted_fields)

    status = "SUCCESS" if confidence >= 0.8 else ("PARTIAL" if confidence >= 0.4 else "FAILED")
    notes = (
        f"Extracted {len([v for v in extracted_fields.values() if v is not None])} fields from {len(files_data)} documents ({unified_filename}). "
        f"Confidence: {confidence:.0%}. "
        + (f"Missing: {', '.join(missing)}." if missing else "All required fields found.")
    )

    # Verify extracted entities against DB (supplier + importer)
    entity_verification = verify_extracted_entities(db, extracted_fields)

    session_id = None
    session_ref = None
    if save_session:
        try:
            session = repo.create_upload_session(
                db=db,
                module_name=module_name,
                filename=unified_filename,
                file_type=primary_file_type,
                file_size_bytes=total_size,
                extraction_status=status,
                confidence_score=confidence,
                extracted_fields=extracted_fields,
                missing_fields=missing,
                extraction_notes=notes,
                created_by=created_by,
            )
            session_id = session.id
            session_ref = session.session_ref
        except Exception as e:
            logger.warning(f"Could not save multi-upload session: {e}")

    return {
        "session_id": session_id,
        "session_ref": session_ref,
        "module_name": module_name,
        "filename": unified_filename,
        "file_type": primary_file_type,
        "file_size_bytes": total_size,
        "extraction_status": status,
        "confidence_score": confidence,
        "extracted_fields": extracted_fields,
        "missing_fields": missing,
        "extraction_notes": notes,
        "raw_text_preview": full_text[:500] if full_text else None,
        **entity_verification,
    }


def parse_raw_text_directly(
    db: Session,
    module_name: str,
    raw_text: str,
    save_session: bool = False,
    created_by: str = "system",
) -> dict:
    """
    Parses raw text directly without file upload (e.g. pasted contact info / address block)
    and returns structured extracted entity or document fields.
    """
    extractor = _EXTRACTOR_MAP.get(module_name)
    if extractor is None:
        extractor = MasterDataEntityExtractor()

    try:
        if isinstance(extractor, MasterDataEntityExtractor):
            extracted_fields = extractor.extract(raw_text, {}, module_name=module_name)
        else:
            extracted_fields = extractor.extract(raw_text, {})
    except Exception as e:
        logger.error(f"Direct raw text extractor error for '{module_name}': {e}")
        extracted_fields = {}

    confidence = extractor.compute_confidence(extracted_fields)
    missing = extractor.missing_required(extracted_fields)
    status = "SUCCESS" if confidence >= 0.8 else ("PARTIAL" if confidence >= 0.4 else "FAILED")

    notes = (
        f"Parsed raw text block ({len(raw_text)} chars). "
        f"Confidence: {confidence:.0%}. "
        + (f"Missing: {', '.join(missing)}." if missing else "All required fields found.")
    )

    # Verify extracted entities against DB (supplier + importer)
    entity_verification = verify_extracted_entities(db, extracted_fields)

    session_id = None
    session_ref = None
    if save_session:
        try:
            session = repo.create_upload_session(
                db=db,
                module_name=module_name,
                filename="pasted_raw_text",
                file_type="text",
                file_size_bytes=len(raw_text.encode("utf-8")),
                extraction_status=status,
                confidence_score=confidence,
                extracted_fields=extracted_fields,
                missing_fields=missing,
                extraction_notes=notes,
                created_by=created_by,
            )
            session_id = session.id
            session_ref = session.session_ref
        except Exception as e:
            logger.warning(f"Could not save text parse session: {e}")

    return {
        "session_id": session_id,
        "session_ref": session_ref,
        "module_name": module_name,
        "filename": "pasted_raw_text",
        "file_type": "text",
        "file_size_bytes": len(raw_text.encode("utf-8")),
        "extraction_status": status,
        "confidence_score": confidence,
        "extracted_fields": extracted_fields,
        "missing_fields": missing,
        "extraction_notes": notes,
        **entity_verification,
    }


def get_upload_sessions_service(db: Session, module_name: str = None, skip: int = 0, limit: int = 50):
    return repo.get_upload_sessions(db, module_name=module_name, skip=skip, limit=limit)

def get_upload_session_by_id_service(db: Session, session_id: int):
    return repo.get_upload_session_by_id(db, session_id)

def soft_delete_upload_session_service(db: Session, session_id: int):
    return repo.soft_delete_upload_session(db, session_id)
