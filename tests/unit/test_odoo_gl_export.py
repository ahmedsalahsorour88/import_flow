import pytest
import io
import csv
import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from database.database import Base, get_db
from main import app
from modules.import_files.model import ImportFile
from modules.financial_settlement.model import LandedCostSettlementRecord
from modules.financial_settlement.schemas import (
    FinancialSettlementCreate,
    ExpenseInvoiceSchema,
    ItemLandedCostSchema,
)
from modules.financial_settlement.service import create_settlement_service
from modules.financial_settlement.odoo_export_service import (
    generate_odoo_journal_entry_service,
    export_odoo_csv_service,
    export_odoo_excel_service,
)

# Test in-memory SQLite DB
SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

@pytest.fixture(scope="function")
def db_session():
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(bind=engine)

@pytest.fixture(scope="function")
def client(db_session):
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    yield TestClient(app)
    app.dependency_overrides.clear()

def _create_sample_settlement(db):
    # 1. Create an Import File
    imp_file = ImportFile(
        import_file_code="IMP-2026-0042",
        company_name="المصرية لمستلزمات الصناعة (EGY Industry)",
        supplier_name="Global Steel & Plastic Co. Ltd",
        project_names="Expansion Project Alexandria 2026",
        current_module="Phase 9",
        current_stage="Settlement",
        progress_percent=90.0,
    )
    db.add(imp_file)
    db.commit()
    db.refresh(imp_file)

    # 2. Create Landed Cost Settlement with all 6 required cost categories:
    # - Freight (Sea)
    # - Clearance (Brokerage)
    # - Local Transport
    # - Customs Duties & Taxes Claim
    # - Demurrage & Storage
    # - Price Adjustment Penalty
    payload = FinancialSettlementCreate(
        import_file_id=imp_file.import_file_id,
        expense_invoices=[
            ExpenseInvoiceSchema(
                invoice_no="INV-SEA-9812",
                category="Freight",
                provider_name="MSC Mediterranean Shipping Co.",
                currency="USD",
                amount_fx=1200.0,
                exchange_rate=50.0, # 60,000 EGP
                amount_egp=60000.0,
                allocation_rule="Value-Based",
            ),
            ExpenseInvoiceSchema(
                invoice_no="CLR-2026-77",
                category="Brokerage",
                provider_name="Al-Ahram Customs Brokerage",
                currency="EGP",
                amount_fx=15000.0,
                exchange_rate=1.0,
                amount_egp=15000.0,
                allocation_rule="Value-Based",
            ),
            ExpenseInvoiceSchema(
                invoice_no="TRK-4401",
                category="Local Transport",
                provider_name="Nile Inland Trucking",
                currency="EGP",
                amount_fx=12000.0,
                exchange_rate=1.0,
                amount_egp=12000.0,
                allocation_rule="Value-Based",
            ),
            ExpenseInvoiceSchema(
                invoice_no="CUSTOMS-DEC-46-991",
                category="Customs Duty",
                provider_name="مصلحة الجمارك المصرية - نافذة",
                currency="EGP",
                amount_fx=85000.0,
                exchange_rate=1.0,
                amount_egp=85000.0,
                allocation_rule="Value-Based",
            ),
            ExpenseInvoiceSchema(
                invoice_no="DEM-MSC-3310",
                category="Demurrage & Storage",
                provider_name="Alexandria Port Authority / MSC",
                currency="USD",
                amount_fx=300.0,
                exchange_rate=50.0, # 15,000 EGP
                amount_egp=15000.0,
                allocation_rule="Value-Based",
            ),
            ExpenseInvoiceSchema(
                invoice_no="PEN-VAL-009",
                category="Price Adjustment / Penalty",
                provider_name="Customs Valuation Authority",
                currency="EGP",
                amount_fx=8000.0,
                exchange_rate=1.0,
                amount_egp=8000.0,
                allocation_rule="Value-Based",
            ),
        ],
        item_landed_costs=[
            ItemLandedCostSchema(
                item_code="RAW-PLAST-01",
                item_name="Industrial Plastic Resins",
                qty=1000,
                gross_weight_kg=20000.0,
                cbm=35.0,
                fob_unit_egp=300.0,
                fob_total_egp=300000.0, # 300,000 EGP
            ),
            ItemLandedCostSchema(
                item_code="RAW-STEEL-02",
                item_name="Cold Rolled Steel Coils",
                qty=500,
                gross_weight_kg=15000.0,
                cbm=12.0,
                fob_unit_egp=400.0,
                fob_total_egp=200000.0, # 200,000 EGP
            ),
        ],
        accountant_name="Mahmoud Accounting",
        notes="Full Landed Cost with all expense components",
    )

    record = create_settlement_service(db, payload)
    return imp_file, record

def test_odoo_journal_entry_balanced_math(db_session):
    """
    Test that generated Odoo journal entry is perfectly balanced (Total Debit == Total Credit).
    """
    imp_file, settlement = _create_sample_settlement(db_session)
    entry = generate_odoo_journal_entry_service(db_session, settlement.settlement_id)

    # Total FOB = 300,000 + 200,000 = 500,000
    # Total Expenses = 60,000 + 15,000 + 12,000 + 85,000 + 15,000 + 8,000 = 195,000
    # Total Landed Cost = 695,000 EGP
    assert entry.total_debit == 695000.0
    assert entry.total_credit == 695000.0
    assert entry.is_balanced is True
    assert entry.difference == 0.0
    assert entry.import_file_code == "IMP-2026-0042"
    assert entry.project_name == "Expansion Project Alexandria 2026"

def test_odoo_journal_lines_category_mapping(db_session):
    """
    Test that each required cost category creates a corresponding credit line with appropriate partner.
    """
    imp_file, settlement = _create_sample_settlement(db_session)
    entry = generate_odoo_journal_entry_service(db_session, settlement.settlement_id)

    # Line 0: Debit Landed Cost
    assert entry.lines[0].debit == 695000.0
    assert entry.lines[0].credit == 0.0
    assert entry.lines[0].cost_category == "Goods"

    # Line 1: Supplier FOB
    assert entry.lines[1].partner_name == "Global Steel & Plastic Co. Ltd"
    assert entry.lines[1].credit == 500000.0

    # Expense Lines categories
    categories = [l.cost_category for l in entry.lines]
    assert "Freight" in categories
    assert "Clearance" in categories
    assert "Transport" in categories
    assert "Customs" in categories
    assert "Demurrage" in categories
    assert "Price_Adjustment" in categories

def test_odoo_csv_export_format(db_session):
    """
    Test Odoo standard CSV export string and headers.
    """
    imp_file, settlement = _create_sample_settlement(db_session)
    csv_str = export_odoo_csv_service(db_session, settlement.settlement_id)

    reader = csv.reader(io.StringIO(csv_str))
    rows = list(reader)
    assert len(rows) > 1

    headers = rows[0]
    assert "line_ids/account_id/code" in headers
    assert "line_ids/debit" in headers
    assert "line_ids/credit" in headers
    assert "line_ids/partner_id/name" in headers

    # Verify first data line has move_id reference
    assert rows[1][0] == f"importflow_settlement_{settlement.settlement_id}"

def test_odoo_excel_export_structure(db_session):
    """
    Test Excel export generates 3 valid sheets with proper accounting voucher structure.
    """
    imp_file, settlement = _create_sample_settlement(db_session)
    excel_bytes = export_odoo_excel_service(db_session, settlement.settlement_id)

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames
    assert "Odoo_GL_Import" in sheet_names
    assert "Accounting_Voucher" in sheet_names
    assert "Items_Landed_Cost" in sheet_names

    ws_voucher = wb["Accounting_Voucher"]
    assert "Import Landed Cost Journal Voucher" in ws_voucher.cell(row=1, column=1).value

def test_odoo_endpoints_rest_api(client, db_session):
    """
    Test FastAPI REST endpoints for Odoo JSON preview, CSV download, and Excel download.
    """
    imp_file, settlement = _create_sample_settlement(db_session)

    # 1. GET JSON
    res_json = client.get(f"/api/v1/financial-settlement/{settlement.settlement_id}/odoo-journal-entry")
    assert res_json.status_code == 200
    data = res_json.json()
    assert data["is_balanced"] is True
    assert data["total_debit"] == 695000.0
    assert len(data["lines"]) == 8 # 1 debit + 1 fob credit + 6 expense credit lines

    # 2. GET CSV
    res_csv = client.get(f"/api/v1/financial-settlement/{settlement.settlement_id}/export-odoo-csv")
    assert res_csv.status_code == 200
    assert "text/csv" in res_csv.headers["content-type"]
    assert "odoo_landed_cost_settlement_" in res_csv.headers["content-disposition"]

    # 3. GET Excel
    res_excel = client.get(f"/api/v1/financial-settlement/{settlement.settlement_id}/export-odoo-excel")
    assert res_excel.status_code == 200
    assert "spreadsheetml.sheet" in res_excel.headers["content-type"]
    assert "accounting_landed_cost_voucher_" in res_excel.headers["content-disposition"]
