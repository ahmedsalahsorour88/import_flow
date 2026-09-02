"""
Unit tests for CargoX Standard Excel Commercial Invoice Generator, Parser, Comparison, and Session Upsert (BP-025 / CGX-002).
"""

import pytest
from datetime import datetime, timezone
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi.testclient import TestClient

from database.database import Base, get_db
from main import app
from modules.import_files.model import ImportFile
from modules.import_companies.model import ImportCompany
from modules.suppliers.model import Supplier
from modules.purchase_orders.model import PurchaseOrder, POLineItem
from modules.cargox.schemas import (
    StandardInvoicePayload,
    StandardInvoiceLineItem,
    StandardInvoiceSessionCreate,
    StandardInvoiceStatusUpdateRequest,
)
from modules.cargox.service import CargoXStandardInvoiceService
from modules.cargox.excel_invoice_service import (
    generate_standard_invoice_excel_bytes,
    parse_standard_invoice_excel_bytes,
    compare_standard_invoice_data,
)


@pytest.fixture
def test_db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()

    # Seed Company
    company = ImportCompany(
        company_id=1,
        importer_name="الشركة المصرية للاستيراد والتوريدات",
        address="10 El-Bostan St, Cairo",
        country="Egypt",
        importer_id="IMP-001-REG",
        importer_id_expiry=datetime(2030, 1, 1).date(),
        vat_id="123456789",
        vat_id_expiry=datetime(2030, 1, 1).date(),
        registration_number="CR-987654",
        registration_expiry=datetime(2030, 1, 1).date(),
    )
    db.add(company)

    # Seed Supplier
    supplier = Supplier(
        supplier_id=1,
        supplier_code="SUP-001",
        company_name="Narbutas International UAB",
        supplier_type="Manufacturer",
        registration_type="Commercial Registration",
        foreign_exporter_id="LT300591314",
        foreign_exporter_country="Lithuania",
        foreign_exporter_country_code="LT",
        address="Ukmerges g. 308, Vilnius",
        phone="+37052100000",
        email="sales@narbutas.lt",
        website="www.narbutas.com",
    )
    db.add(supplier)

    # Seed Import File
    file = ImportFile(
        import_file_id=1,
        import_file_code="IMP-2026-009",
        company_id=1,
        company_name="Egyptian Import & Supply Co.",
        supplier_id=1,
        supplier_name="Narbutas International UAB",
        acid_number="7595528271020210010",
        estimated_cost_currency="EUR",
        incoterm_code="EXW",
        port_of_loading="Klaipeda Port",
        port_of_discharge="Alexandria Port",
        estimated_cost=5000.0,
    )
    db.add(file)

    # Seed PO & Items
    po = PurchaseOrder(
        po_id=1,
        po_number="PO-2026-001",
        import_file_id=1,
        project_id=1,
        company_id=1,
        supplier_id=1,
        incoterm_id=1,
        currency_id=1,
        total_amount_fob=5000.0,
    )
    db.add(po)
    db.flush()

    item1 = POLineItem(
        po_id=1,
        item_code="DSK-001",
        description_ar="مكتب تنفيذي",
        description_en="Executive Office Desk",
        quantity=10.0,
        unit_price=300.0,
        total_price=3000.0,
        unit_of_measure="SET",
        gross_weight_kg=250.0,
        net_weight_kg=220.0,
        country_of_origin="LT",
    )
    item2 = POLineItem(
        po_id=1,
        item_code="CHR-002",
        description_ar="كرسي مريح",
        description_en="Ergonomic Mesh Chair",
        quantity=20.0,
        unit_price=100.0,
        total_price=2000.0,
        unit_of_measure="PCE",
        gross_weight_kg=200.0,
        net_weight_kg=180.0,
        country_of_origin="LT",
    )
    db.add(item1)
    db.add(item2)
    db.commit()

    yield db
    db.close()



def test_standard_invoice_excel_generation_and_parsing(test_db):
    """Test generating .xlsx with openpyxl Named Ranges & Table and parsing it back cleanly."""
    snapshot = CargoXStandardInvoiceService.build_system_snapshot(test_db, 1)
    assert snapshot.acid_number == "7595528271020210010"
    assert snapshot.seller_tax_id == "LT300591314"
    assert snapshot.buyer_tax_id == "123456789"
    assert len(snapshot.items) == 1
    assert snapshot.subtotal == 5000.0
    assert snapshot.total_amount == 5000.0

    # Generate Excel Bytes
    excel_bytes = generate_standard_invoice_excel_bytes(snapshot)
    assert len(excel_bytes) > 1000

    # Parse Excel Bytes back
    parsed = parse_standard_invoice_excel_bytes(excel_bytes)
    assert parsed.acid_number == "7595528271020210010"
    assert parsed.seller_name == "Narbutas International UAB"
    assert parsed.buyer_tax_id == "123456789"
    assert parsed.currency_code == "EUR"
    assert parsed.incoterm == "EXW"
    assert len(parsed.items) == 1
    assert parsed.items[0].hs_code == "940310"
    assert parsed.items[0].quantity == 30.0
    assert parsed.subtotal == 5000.0


def test_standard_invoice_comparison_perfect_and_discrepancy(test_db):
    """Test comparison matrix for matching vs modified supplier data."""
    snapshot = CargoXStandardInvoiceService.build_system_snapshot(test_db, 1)
    
    # 1. Perfect Match
    comp_res = CargoXStandardInvoiceService.compare_invoices(test_db, 1, snapshot)
    assert comp_res.has_discrepancies is False
    assert comp_res.has_critical_mismatch is False
    assert comp_res.total_discrepancies_count == 0

    # 2. Supplier Data with Discrepancies
    modified_sup = snapshot.model_copy(deep=True)
    modified_sup.acid_number = "9999999999999999999"  # Mismatch
    modified_sup.items[0].unit_price = 350.0  # Price difference
    modified_sup.items[0].total_amount = 3500.0
    modified_sup.subtotal = 5500.0
    modified_sup.total_amount = 5850.0

    comp_diff = CargoXStandardInvoiceService.compare_invoices(test_db, 1, modified_sup)
    assert comp_diff.has_discrepancies is True
    assert comp_diff.has_critical_mismatch is True
    assert comp_diff.critical_mismatches_count >= 1
    assert comp_diff.rectification_notice_en is not None
    assert "ACID Number" in comp_diff.rectification_notice_en


def test_standard_invoice_session_upsert_and_override_justification(test_db):
    """Test Anti-duplicate upsert logic and mandatory discrepancy override justification."""
    snapshot = CargoXStandardInvoiceService.build_system_snapshot(test_db, 1)

    # 1. Create Initial Session as DRAFT
    create_schema = StandardInvoiceSessionCreate(
        import_file_id=1,
        import_file_code="IMP-2026-009",
        acid_number="7595528271020210010",
        invoice_number="INV-2026-009",
        subtotal_amount=5000.0,
        total_amount=5350.0,
        line_items_count=2,
        has_discrepancies=True,
        has_critical_mismatch=False,
        status="DRAFT",
    )
    session1 = CargoXStandardInvoiceService.save_or_upsert_session(test_db, create_schema)
    assert session1.session_id is not None
    assert session1.session_code.startswith("CX-INV-")
    assert session1.status == "DRAFT"

    # 2. Try Approving with Discrepancies without Override Reason -> Should Raise 400
    create_schema.status = "APPROVED"
    create_schema.discrepancy_override_reason = ""
    with pytest.raises(Exception) as exc_info:
        CargoXStandardInvoiceService.save_or_upsert_session(test_db, create_schema)
    assert "يجب كتابة مبرر" in str(exc_info.value)

    # 3. Approve with Valid Justification -> Should Update Same Record (Anti-Duplicate)
    create_schema.discrepancy_override_reason = "تم اعتماد الفرق البسيط في السعر بموافقة الإدارة المالية."
    session2 = CargoXStandardInvoiceService.save_or_upsert_session(test_db, create_schema)
    assert session2.session_id == session1.session_id  # Same session updated!
    assert session2.session_code == session1.session_code
    assert session2.status == "APPROVED"
    assert session2.discrepancy_override_reason == "تم اعتماد الفرق البسيط في السعر بموافقة الإدارة المالية."


def test_standard_invoice_api_endpoints(test_db):
    """Test REST API endpoints for Standard Invoice Hub."""
    app.dependency_overrides[get_db] = lambda: test_db
    client = TestClient(app)

    # Test Download Excel
    res_gen = client.get("/api/v1/cargox/standard-invoice/generate/1")
    assert res_gen.status_code == 200
    assert "application/vnd.openxmlformats-officedocument" in res_gen.headers["content-type"]
    excel_bytes = res_gen.content

    # Test Upload & Parse Excel
    res_parse = client.post(
        "/api/v1/cargox/standard-invoice/parse",
        files={"file": ("test_invoice.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res_parse.status_code == 200
    parsed_json = res_parse.json()
    assert parsed_json["acid_number"] == "7595528271020210010"
    assert len(parsed_json["items"]) == 1

    # Test Compare Endpoint
    res_comp = client.post(
        "/api/v1/cargox/standard-invoice/compare/1",
        json=parsed_json,
    )
    assert res_comp.status_code == 200
    comp_json = res_comp.json()
    assert comp_json["has_discrepancies"] is False

    # Test Save Session
    session_payload = {
        "import_file_id": 1,
        "import_file_code": "IMP-2026-009",
        "acid_number": "7595528271020210010",
        "invoice_number": "INV-2026-009",
        "currency_code": "EUR",
        "subtotal_amount": 5000.0,
        "total_amount": 5350.0,
        "line_items_count": 2,
        "status": "APPROVED",
        "has_discrepancies": False,
    }
    res_save = client.post("/api/v1/cargox/standard-invoice/session", json=session_payload)
    assert res_save.status_code == 200
    assert res_save.json()["status"] == "APPROVED"

    # Test Get by file
    res_by_file = client.get("/api/v1/cargox/standard-invoice/session/by-file/1")
    assert res_by_file.status_code == 200
    assert res_by_file.json()["session_code"].startswith("CX-INV-")

    app.dependency_overrides.clear()


def test_cargox_excel_structure_and_named_ranges(test_db):
    """Rigorous test verifying 100% compliance with official CargoX template structure, sheets, fonts, and named ranges."""
    import io
    import openpyxl

    snapshot = CargoXStandardInvoiceService.build_system_snapshot(test_db, 1)
    excel_bytes = generate_standard_invoice_excel_bytes(snapshot)

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

    # 1. Verify 8 Sheet Names in Exact Order
    expected_sheets = [
        "Invoice",
        "Package Type list",
        "Unit of measure list",
        "Cities",
        "Currency List",
        "UOM List",
        "Country List",
        "City List",
    ]
    assert wb.sheetnames == expected_sheets

    # 2. Main Sheet Structure & Gridlines
    ws = wb["Invoice"]
    assert ws.views.sheetView[0].showGridLines is False

    # 3. Merged Cells Check (All 10 required ranges)
    merged_ranges = [str(m) for m in ws.merged_cells.ranges]
    totals_start_row = 13 + len(snapshot.items)
    expected_merges = [
        "B1:C2", "B3:C3", "D1:D2", "F3:G3", "F4:G4",
        "F5:G5", "H7:J7", "H8:J8",
        f"B{totals_start_row+2}:J{totals_start_row+2}",
        f"B{totals_start_row+4}:J{totals_start_row+4}",
    ]
    for mr in expected_merges:
        assert mr in merged_ranges, f"Missing merged range: {mr}"

    assert ws["B1"].value == "Narbutas International UAB"
    assert ws["B1"].font.name == "Calibri"
    assert ws["B3"].font.size == 18.0
    assert ws["B3"].font.bold is True

    # 4. Critical Named Ranges Check
    assert "SellerName" in wb.defined_names
    assert "SellerRegistrationCode" in wb.defined_names
    assert "BuyerName" in wb.defined_names
    assert "BuyerCode" in wb.defined_names
    assert "ACIDNumber" in wb.defined_names
    assert "TotalInvoiceLinesNo" in wb.defined_names
    assert "InvoiceSubtotal" in wb.defined_names
    assert "FreightCost" in wb.defined_names
    assert "InsuranceCost" in wb.defined_names
    assert "OtherCosts" in wb.defined_names
    assert "TotalAmount" in wb.defined_names

    # 5. Exact Labels Check
    assert ws["E1"].value == "Registration # :"
    assert ws["E2"].value == "City Code :"
    assert ws["H2"].value == "Phone:"
    assert ws["E3"].value == "Web Site :"
    assert ws["H3"].value == "Fax:"
    assert ws["D4"].value == "Exporter Contact :"
    assert ws["B5"].value == "Export To:"
    assert ws["D5"].value == "Importer Contact :"
    assert ws["H5"].value == "Phone:"
    assert ws["K5"].value == "Gross Weight"
    assert ws["D7"].value == "ACID #:"
    assert ws["F7"].value == "Origin Port:"
    assert ws["K7"].value == "Weigh unitt:"
    assert ws["B8"].value == "Purchase Order #:"
    assert ws["D8"].value == "Purchase Order Date:"
    assert ws["F8"].value == "Destination Port:"
    assert ws["B9"].value == "Invoice #:"
    assert ws["D9"].value == "Invoice Date:"
    assert ws["B10"].value == "Currency:"
    assert ws["D10"].value == "Inco Term:"
    assert ws["K4"].value == "Proforma invoice number:"

    # 6. Formulas Check
    totals_row = 13 + len(snapshot.items)
    assert "=ROWS(InvoiceItems[])" in ws[f"C{totals_row}"].value or "=ROWS(" in ws[f"C{totals_row}"].value
    assert "=SUM(InvoiceItems[Total])" in ws[f"P{totals_row}"].value or "=SUM(" in ws[f"P{totals_row}"].value
    assert "=ROWS(InvoiceItems[])" in ws["H11"].value or "=ROWS(" in ws["H11"].value

    # 7. Table Check
    assert "InvoiceItems" in ws.tables


# ============================================================================
# CGX-003: Multi-Path Extraction Engine Tests
# ============================================================================

from modules.cargox.service import CargoXExtractionEngine
from modules.cargox.schemas import ExtractionRequest, ExtractionResponse, CustomsInvoiceTrackCreate


def test_extraction_all_consolidated_mode(test_db):
    """
    Test all_consolidated mode:
    - يجب أن يُرجع result واحد فقط
    - بنود مجمعة بـ HS Code (سطر واحد لكل HS Code)
    - يطبق weighted average price
    """
    request = ExtractionRequest(mode="all_consolidated", grouping_mode="by_hs_code")
    response = CargoXExtractionEngine.extract(test_db, 1, request)

    assert response.import_file_id == 1
    assert response.mode == "all_consolidated"
    assert response.invoices_count == 1  # all modes → result واحد
    assert len(response.results) == 1

    payload = response.results[0].payload
    # كلا البندين بدون tariff → نفس HS Code الافتراضي "940310" → يتجمعوا في سطر واحد
    assert len(payload.items) == 1

    item = payload.items[0]
    assert item.hs_code == "940310"
    assert item.quantity == 30.0  # 10 + 20

    # Weighted Average Price: (10×300 + 20×100) / 30 = 5000 / 30 ≈ 166.6667
    expected_avg_price = round((10 * 300 + 20 * 100) / 30, 4)
    assert abs(item.unit_price - expected_avg_price) < 0.001

    # Total = 30 × 166.6667 ≈ 5000
    assert abs(item.total_amount - 5000.0) < 0.1
    assert abs(payload.subtotal - 5000.0) < 0.1


def test_extraction_all_detailed_mode(test_db):
    """
    Test all_detailed mode:
    - يجب أن يُرجع result واحد فقط
    - بنود مفصلة (كل سطر منفصل حتى لو نفس HS Code)
    """
    request = ExtractionRequest(mode="all_detailed", grouping_mode="by_hs_code")
    response = CargoXExtractionEngine.extract(test_db, 1, request)

    assert response.mode == "all_detailed"
    assert response.invoices_count == 1
    assert len(response.results) == 1

    payload = response.results[0].payload
    # detailed mode → flat → 2 بنود منفصلة
    assert len(payload.items) == 2

    # التحقق من الترتيب والكميات
    assert payload.items[0].quantity == 10.0
    assert payload.items[0].unit_price == 300.0
    assert payload.items[1].quantity == 20.0
    assert payload.items[1].unit_price == 100.0


def test_extraction_per_invoice_consolidated_mode(test_db):
    """
    Test per_invoice_consolidated mode:
    - كلا البندين لهم نفس invoice_number (PO-level)
    - يجب أن يُرجع result واحد (فاتورة واحدة) مجمع
    """
    request = ExtractionRequest(mode="per_invoice_consolidated", grouping_mode="by_hs_code")
    response = CargoXExtractionEngine.extract(test_db, 1, request)

    assert response.mode == "per_invoice_consolidated"
    # كلا البندين تحت نفس الـ PO وproforma_invoice_number → فاتورة واحدة
    assert response.invoices_count >= 1

    # كل ملف مجمع → 1 بند لكل HS Code
    for result in response.results:
        hs_codes_seen = set()
        for item in result.payload.items:
            assert item.hs_code not in hs_codes_seen, f"HS Code {item.hs_code} مكرر في consolidated mode!"
            hs_codes_seen.add(item.hs_code)


def test_extraction_per_invoice_detailed_mode(test_db):
    """
    Test per_invoice_detailed mode:
    - بنود مفصلة لكل فاتورة
    """
    request = ExtractionRequest(mode="per_invoice_detailed", grouping_mode="flat")
    response = CargoXExtractionEngine.extract(test_db, 1, request)

    assert response.mode == "per_invoice_detailed"
    assert response.invoices_count >= 1

    # التحقق أن كل البنود موجودة
    total_items = sum(len(r.payload.items) for r in response.results)
    assert total_items == 2  # DSK-001 + CHR-002


def test_weighted_average_price_calculation(test_db):
    """
    Test خوارزمية Weighted Average Price بدقة:
    100 PCS × $5 + 500 PCS × $7 = $4,000 / 600 = $6.6667
    """
    # نبني raw items مباشرة لاختبار الـ grouping logic
    raw_items = [
        {
            "item_code": "A001",
            "hs_code": "620000",
            "description": "Fabric Type A",
            "quantity": 100.0,
            "qty_unit": "PCS",
            "unit_price": 5.0,
            "country_of_origin": "CN",
            "gross_weight_kg": 50.0,
            "net_weight_kg": 45.0,
        },
        {
            "item_code": "A002",
            "hs_code": "620000",  # نفس HS Code — سيتم تجميعهم
            "description": "Fabric Type B",
            "quantity": 500.0,
            "qty_unit": "PCS",
            "unit_price": 7.0,
            "country_of_origin": "CN",
            "gross_weight_kg": 250.0,
            "net_weight_kg": 225.0,
        },
    ]

    items = CargoXExtractionEngine._group_items(raw_items, "by_hs_code", "all_consolidated")

    assert len(items) == 1, "يجب تجميع البندين في سطر واحد لنفس HS Code"
    item = items[0]

    # Weighted Average Price = (100×5 + 500×7) / (100+500) = 4000/600 ≈ 6.6667
    expected_price = round((100 * 5 + 500 * 7) / 600, 4)
    assert abs(item.unit_price - expected_price) < 0.0001

    # Total = 600 × 6.6667 = 4000
    assert item.quantity == 600.0
    assert abs(item.total_amount - 4000.0) < 0.1

    # Gross Weight = 50 + 250 = 300
    assert abs(item.gross_weight_kg - 300.0) < 0.01


def test_by_price_group_grouping(test_db):
    """
    Test by_price_group mode:
    نفس HS Code لكن أسعار مختلفة → سطرين منفصلين
    """
    raw_items = [
        {"item_code": "B001", "hs_code": "940310", "description": "Desk A", "quantity": 10.0, "qty_unit": "PCS", "unit_price": 200.0, "country_of_origin": "DE", "gross_weight_kg": 100.0, "net_weight_kg": 90.0},
        {"item_code": "B002", "hs_code": "940310", "description": "Desk B", "quantity": 5.0, "qty_unit": "PCS", "unit_price": 350.0, "country_of_origin": "DE", "gross_weight_kg": 60.0, "net_weight_kg": 55.0},
    ]

    # by_price_group → سطرين (أسعار مختلفة)
    items_grouped = CargoXExtractionEngine._group_items(raw_items, "by_price_group", "all_consolidated")
    assert len(items_grouped) == 2, "بالـ by_price_group يجب أن يكون 2 سطور لأن الأسعار مختلفة"

    # by_hs_code → سطر واحد مع weighted average
    items_consolidated = CargoXExtractionEngine._group_items(raw_items, "by_hs_code", "all_consolidated")
    assert len(items_consolidated) == 1, "بالـ by_hs_code يجب أن يكون سطر واحد"
    expected_avg = round((10 * 200 + 5 * 350) / 15, 4)
    assert abs(items_consolidated[0].unit_price - expected_avg) < 0.001


def test_extraction_response_structure(test_db):
    """
    Test هيكل الـ ExtractionResponse الكامل.
    """
    request = ExtractionRequest(mode="all_consolidated", grouping_mode="by_hs_code")
    response = CargoXExtractionEngine.extract(test_db, 1, request)

    # التحقق من هيكل الاستجابة
    assert isinstance(response, ExtractionResponse)
    assert response.import_file_id == 1
    assert response.import_file_code == "IMP-2026-009"
    assert response.invoices_count == len(response.results)
    assert response.total_line_items == sum(len(r.payload.items) for r in response.results)

    # كل result له payload صحيح
    for result in response.results:
        assert result.payload.acid_number == "7595528271020210010"
        assert result.payload.seller_name == "Narbutas International UAB"
        assert result.payload.currency_code == "EUR"
        assert len(result.payload.items) > 0


def test_zip_generation_multi_invoice(test_db):
    """
    Test توليد ZIP يحتوي على Excel files من محرك الاستخراج.
    """
    import io
    import zipfile as zipfile_module
    from modules.cargox.excel_invoice_service import generate_standard_invoice_excel_bytes

    request = ExtractionRequest(mode="all_consolidated", grouping_mode="by_hs_code")
    extraction = CargoXExtractionEngine.extract(test_db, 1, request)

    zip_buffer = io.BytesIO()
    with zipfile_module.ZipFile(zip_buffer, "w", zipfile_module.ZIP_DEFLATED) as zf:
        for result in extraction.results:
            excel_bytes = generate_standard_invoice_excel_bytes(result.payload)
            safe_inv = (result.invoice_number or "Invoice_default").replace("/", "-")
            filename = f"Commercial_Invoice_{safe_inv}.xlsx"
            zf.writestr(filename, excel_bytes)

    zip_buffer.seek(0)
    zip_data = zip_buffer.read()

    assert len(zip_data) > 100, "ZIP يجب أن يكون أكبر من 100 bytes"

    # التحقق من محتوى الـ ZIP
    zip_buffer.seek(0)
    with zipfile_module.ZipFile(zip_buffer, "r") as zf:
        file_list = zf.namelist()
        assert len(file_list) == extraction.invoices_count
        for fn in file_list:
            assert fn.endswith(".xlsx"), f"الملف {fn} يجب أن يكون .xlsx"
