import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


class MasterDataExportImportHelper:
    @staticmethod
    def create_excel_template(columns: List[str], sample_row: Dict[str, Any] = None) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Template"

        # Headers styling
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")

        ws.append(columns)
        for col_num, _ in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        if sample_row:
            row_vals = [sample_row.get(col, "") for col in columns]
            ws.append(row_vals)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def parse_excel_file(file_bytes: bytes, required_columns: List[str]) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        
        parsed_data = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for idx, h in enumerate(headers):
                if h and idx < len(row):
                    val = row[idx]
                    row_dict[h] = str(val).strip() if val is not None else ""
            parsed_data.append(row_dict)

        return parsed_data

    @staticmethod
    def export_to_excel(filename_title: str, headers: List[str], rows_data: List[List[Any]]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = filename_title[:30]

        # Header Row
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        for r_idx, r_data in enumerate(rows_data, start=2):
            ws.append([str(v) if v is not None else "" for v in r_data])
            fill_color = "F8F9FA" if r_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for c_num in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_num)
                cell.fill = row_fill
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def export_to_pdf(title: str, headers: List[str], rows_data: List[List[Any]]) -> bytes:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=16,
            leading=20,
            alignment=1, # Center
            textColor=colors.HexColor("#2C3E50"),
        )
        cell_style = ParagraphStyle(
            name="CellStyle",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
        )
        header_cell_style = ParagraphStyle(
            name="HeaderCellStyle",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )

        elements = []
        elements.append(Paragraph(f"<b>ImportFlow ERP - {title}</b>", title_style))
        elements.append(Spacer(1, 15))

        table_data = []
        # Header Row
        table_data.append([Paragraph(h, header_cell_style) for h in headers])

        # Data Rows
        for row in rows_data:
            table_data.append([Paragraph(str(v) if v is not None else "", cell_style) for v in row])

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ]))

        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return buf.getvalue()
